import string
import datetime
import sys
from pdftables import get_tables
import openpyxl
from openpyxl.cell import get_column_letter
import glob
import os
import warnings
import timeit
warnings.filterwarnings("ignore")

def fill(wb,matrix,title,n,postmonth,postday):
	wb.create_sheet(title = title)
	sheet = wb.get_sheet_by_name(title)
	flag1=flag2=0
	res = get_tables(matrix)
	sheet['A1']="Label Index"
	sheet['B1']="Date"
	sheet['C1']="Low Price"
	sheet['D1']="High Price"
	sheet['E1']="Average"
	sheet['F1']="Unit"
	sheet['G1']="Frequency"
	if n == 14:
		n=-2
	elif n == 7:
		n=-3
	else:
		n=-4
	k=0
	for subres in res:
		i=2
		for subsubres in subres:
			if 'Med cfr' in subsubres[0]:
				flag1=1
			if ('*' in subsubres[0] or (subsubres[0]=="" and subsubres[2]!="")) and flag1 == 1:
				flag2=1
			if flag1 == 1 and subsubres[1]!="" and flag2 == 0:
				prod_list=["Sulphur CFR Med (incl N. Africa)","Sulphur CFR Med (small lots N africa)","Sulphur CFR Med (small lots Others Markets)","Sulphur CFR North Africa (contract)","Sulphur FOB Med (small lots other markets)","Sulphur CFR China (Contract)","Sulphur CFR China (Spot)","Sulphur CFR India (Spot)","Sulphur (Liquid) CFR Brazil","Sulphur FOB Vancouver (Contract)","Sulphur FOB Vancouver (Spot)","Sulphur FOB California (Spot)","Sulphur FOB Middle East","Sulphur FOB Middle East (Contract)","Sulphur FOB Middle East (Spot)","Sulphur FOB Qatar (Tasweeq QSP)","Sulphur FOB Saudi Arabia (Armaco)","Sulphur FOB Middle East (Adnoc)","Sulphur CPT NW Europe","Sulphur DEL Benelux","Sulphur CFR Tampa/Central Florida Deliv.","Sulphur CFR Houston (Spot)","Sulphur Ex-tank Galveston"]
				if i == len(prod_list) + 2:
					print "Done"
				else:
					sheet['A'+str(i)]=prod_list[i-2]
					sheet['B'+str(i)]=(str(postday)+"/"+str(postmonth)+"/2016").strip("'")
					if "-" in subsubres[n]:
						sheet['C'+str(i)]=float(subsubres[n].split('-')[0].strip("'").strip('*'))
						sheet['D'+str(i)]=float(subsubres[n].split('-')[1].strip("'").strip('*'))
					else:
						sheet['C'+str(i)]=sheet['D'+str(i)]=float(subsubres[n].strip('*').strip("'"))
					sheet['E'+str(i)]="=AVERAGE(C"+str(i)+":D"+str(i)+")"
					sheet['F'+str(i)]="$/T"
					sheet['G'+str(i)]="Weekly"
					i+=1


def main(argv):
	wb = openpyxl.Workbook()
	wb.remove_sheet(wb.active)
	start = timeit.default_timer()
	i=0
	j = len(glob.glob(sys.argv[1]+"/*.pdf"))
	for pdf_name in sorted(glob.glob(sys.argv[1]+"/*.pdf")):
		matrix = open(pdf_name, "rb")
		day=pdf_name.split("/")[1].split(".")[0].split("_")[-1].split("-")[2]
		month=pdf_name.split("/")[1].split(".")[0].split("_")[-1].split("-")[1]
		print "Parsing > "+day+" "+month 
		if i!= j-1:
			post = datetime.datetime(2016,int(month),int(day)) - datetime.timedelta(days = 14)
			month = ["Janvier","Fevrier","Mars","Avril","Mai","Juin","Juillet","Aout","Septembre","Octobre","Novembre","Decembre"][post.month-1]
			day = post.day
			title=str(day)+" "+month
			fill(wb,matrix,title,14,post.month,post.day)
		else:
			post = datetime.datetime(2016,int(month),int(day)) - datetime.timedelta(days = 14)
			month = ["Janvier","Fevrier","Mars","Avril","Mai","Juin","Juillet","Aout","Septembre","Octobre","Novembre","Decembre"][post.month-1]
			day = post.day
			title=str(day)+" "+month
			fill(wb,matrix,title,14,post.month,post.day)
			day=pdf_name.split("/")[1].split(".")[0].split("_")[-1].split("-")[2]
			month=pdf_name.split("/")[1].split(".")[0].split("_")[-1].split("-")[1]
			post = datetime.datetime(2016,int(month),int(day)) - datetime.timedelta(days = 7)
			month = ["Janvier","Fevrier","Mars","Avril","Mai","Juin","Juillet","Aout","Septembre","Octobre","Novembre","Decembre"][post.month-1]
			day = post.day
			title=str(day)+" "+month
			fill(wb,matrix,title,7,post.month,post.day)
			day=pdf_name.split("/")[1].split(".")[0].split("_")[-1].split("-")[2]
			month=pdf_name.split("/")[1].split(".")[0].split("_")[-1].split("-")[1]
			post = datetime.datetime(2016,int(month),int(day))
			month = ["Janvier","Fevrier","Mars","Avril","Mai","Juin","Juillet","Aout","Septembre","Octobre","Novembre","Decembre"][post.month-1]
			day = post.day
			title=str(day)+" "+month
			fill(wb,matrix,title,0,post.month,post.day)
		i+=1
	wb.save("out.xlsx")
	stop = timeit.default_timer()
	print stop - start

if __name__ == '__main__':
    main(sys.argv)
