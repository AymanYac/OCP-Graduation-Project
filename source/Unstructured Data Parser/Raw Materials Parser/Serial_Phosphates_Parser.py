import string
import datetime
import sys
from pdftables import get_tables
import openpyxl
from openpyxl.cell import get_column_letter
import glob
import os
import warnings
import timeit
warnings.filterwarnings("ignore")

def echo(wb,matrix,title,n,postmonth,postday):
	wb.create_sheet(title = title)
	sheet = wb.get_sheet_by_name(title)
	res = get_tables(matrix)
	for subres in res:
		for subsubres in subres:
			for elem in subsubres:
				sheet[get_column_letter(subsubres.index(elem)+1)+str(subres.index(subsubres)+1)]=elem

def fill(wb,matrix,title,n,postmonth,postday):
	#echo(wb,matrix,title,n,postmonth,postday)
	#return
	wb.create_sheet(title = title)
	sheet = wb.get_sheet_by_name(title)
	flag1=flag2=0
	res = get_tables(matrix)
	sheet['A1']="Label Index"
	sheet['B1']="Date"
	sheet['C1']="Low Price"
	sheet['D1']="High Price"
	sheet['E1']="Average"
	sheet['F1']="Unit"
	sheet['G1']="Frequency"
	if title.split(" ")[1] not in ["Mai","Juin","Juillet","Aout"] or title == "2 Juin":
		if n == 14:
			n=-2
		elif n == 7:
			n=-3
		else:
			n=-4
	else:
		if n == 14:
			n=-1
		elif n == 7:
			n=-2
		else:
			n=-3
	k=0
	print n
	l=[]
	for subres in res:
		m=[]
		for subsubres in subres:
			v=[]
			for item in subsubres:
				if item=="" and subsubres.index(item)!=0:
					v.append(subsubres[subsubres.index(item)-1])
				else:
					v.append(item)
			m.append(v)
		l.append(m)

	for subres in l:
		i=2
		for subsubres in subres:
			fla=0
			if 'US Gulf fob bulk' in subsubres[0]:
				flag1=1
			if 'Casablanca' in subsubres[0]:
				flag2=1
			if flag1 == 1 and subsubres[0] not in ["PHOSPHATE ROCK","PHOSPHORIC ACID","GTSP","MAP","Black Sea fob bulk",""] and flag2 == 0:
				prod_list=["DAP FOB US Gulf","DAP FOB Morocco","DAP FOB Tunisia","DAP FOB Jordan","DAP FOB Saudi Arabia","DAP FOB Baltic","DAP FOB Australia","DAP FOB China","DAP FCA Benelux","DAP bulk CFR India","DAP FOB Nola","DAP FOB C.Florida","TSP FOB Bulgaria","TSP FOB Morocco","TSP FOB Tunisia","TSP FOB Lebanon","TSP FOB Mexico","TSP Bagged FOB China","MAP FOB Black Sea","MAP FOB Baltic","MAP CFR Brazil","Phosphoric Acid FOB US Golf","Phosphoric Acid CFR India","Phosphate Rock FOB Casablanca"]
				if i == len(prod_list) + 2:
					print "Done"
				else:
					sheet['A'+str(i)]=prod_list[i-2]
					sheet['B'+str(i)]=(str(postday)+"/"+str(postmonth)+"/2016").strip("'")
					if "-" in subsubres[n]:
						sheet['C'+str(i)]=float(subsubres[n].split('-')[0].strip("'").strip('*'))
						sheet['D'+str(i)]=float(subsubres[n].split('-')[1].strip("'").strip('*'))
					else:
						if "n.m" not in subsubres[n]:
							sheet['C'+str(i)]=sheet['D'+str(i)]=subsubres[n].strip('*').strip("'")
						else:
							sheet['C'+str(i)]=sheet['D'+str(i)]="n.m."
							fla=1
					if fla == 0:
						sheet['E'+str(i)]="=AVERAGE(C"+str(i)+":D"+str(i)+")"
					else:
						sheet['E'+str(i)]="n.m."
					sheet['F'+str(i)]="$/T"
					sheet['G'+str(i)]="Weekly"
					i+=1


def main(argv):
	wb = openpyxl.Workbook()
	wb.remove_sheet(wb.active)
	start = timeit.default_timer()
	i=0
	j = len(glob.glob(sys.argv[1]+"/*.pdf"))
	for pdf_name in sorted(glob.glob(sys.argv[1]+"/*.pdf")):
		matrix = open(pdf_name, "rb")
		day=pdf_name.split("/")[1].split(".")[0].split("_")[-1].split("-")[2]
		month=pdf_name.split("/")[1].split(".")[0].split("_")[-1].split("-")[1]
		print "Parsing > "+day+" "+month 
		if i!= j-1:
			post = datetime.datetime(2016,int(month),int(day)) - datetime.timedelta(days = 14)
			month = ["Janvier","Fevrier","Mars","Avril","Mai","Juin","Juillet","Aout","Septembre","Octobre","Novembre","Decembre"][post.month-1]
			day = post.day
			title=str(day)+" "+month
			fill(wb,matrix,title,14,post.month,post.day)
		else:
			post = datetime.datetime(2016,int(month),int(day)) - datetime.timedelta(days = 14)
			month = ["Janvier","Fevrier","Mars","Avril","Mai","Juin","Juillet","Aout","Septembre","Octobre","Novembre","Decembre"][post.month-1]
			day = post.day
			title=str(day)+" "+month
			fill(wb,matrix,title,14,post.month,post.day)
			day=pdf_name.split("/")[1].split(".")[0].split("_")[-1].split("-")[2]
			month=pdf_name.split("/")[1].split(".")[0].split("_")[-1].split("-")[1]
			post = datetime.datetime(2016,int(month),int(day)) - datetime.timedelta(days = 7)
			month = ["Janvier","Fevrier","Mars","Avril","Mai","Juin","Juillet","Aout","Septembre","Octobre","Novembre","Decembre"][post.month-1]
			day = post.day
			title=str(day)+" "+month
			fill(wb,matrix,title,7,post.month,post.day)
			day=pdf_name.split("/")[1].split(".")[0].split("_")[-1].split("-")[2]
			month=pdf_name.split("/")[1].split(".")[0].split("_")[-1].split("-")[1]
			post = datetime.datetime(2016,int(month),int(day))
			month = ["Janvier","Fevrier","Mars","Avril","Mai","Juin","Juillet","Aout","Septembre","Octobre","Novembre","Decembre"][post.month-1]
			day = post.day
			title=str(day)+" "+month
			fill(wb,matrix,title,0,post.month,post.day)
		i+=1
	wb.save("out.xlsx")
	stop = timeit.default_timer()
	print stop - start

if __name__ == '__main__':
    main(sys.argv)
