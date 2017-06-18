import openpyxl
from openpyxl.cell import get_column_letter
import sys
import glob
import os
import re
import csv
import warnings
warnings.filterwarnings("ignore")

dapPtrn = re.compile(".*DAP.*")
mapPtrn = re.compile(".*MAP.*")
mgaPtrn = re.compile(".*Acid.*")
tspPtrn = re.compile(".*TSP.*")
rckPtrn = re.compile(".*Rock.*")
expPatern = re.compile(".*xport.*$")
expPatern2 = re.compile(".*rade.*$")
eorPatern = re.compile("PIT.*")
expRowPatern = re.compile(".*20[0-9][0-9].*")
expColPatern = re.compile(".*ountries.*")
impRowPatern = re.compile(".*ountries.*")
eicPatern = re.compile(".*ORLD.*")
eicPatern2 = re.compile(".*orld.*")
subPatern = re.compile(".*otal.*")

for classeur in glob.glob(sys.argv[1]+"/*.xlsx"):
    name = classeur.split("\\")[-1]
    fields = name.split("_")
    if len(fields) not in [6,4]:
        print "Classeur "+name+" non pris en charge : nom de fichier non proprement formatte"
        break
    if len(fields) == 4:
    	prod=fields[0]
    	year = fields[-1].split(".")[0]
    	quart = "Q4"
    	typ = "ANN"
    else:	
	    prod=fields[0]
	    year = fields[-3]
	    quart = fields[-2]
	    typ = fields[-1].split(".")[0]
    if prod=="PHOSPHATE":
        if not os.path.exists(year+"/"+quart+"/ROCK_"+typ):
            os.makedirs(year+"/"+quart+"/ROCK_"+typ+"/EXPORTS")
            os.makedirs(year+"/"+quart+"/ROCK_"+typ+"/PROD")
        else:
            print year+"/"+quart+"_ROCK_"+typ+" exists, back up existing data, clean directory and try again"
            continue
    if prod!="PHOSPHATE":
        if not os.path.exists(year+"/"+quart+"/DAP_"+typ):
            os.makedirs(year+"/"+quart+"/DAP_"+typ+"/EXPORTS")
            os.makedirs(year+"/"+quart+"/MAP_"+typ+"/EXPORTS")
            os.makedirs(year+"/"+quart+"/MGA_"+typ+"/EXPORTS")
            os.makedirs(year+"/"+quart+"/TSP_"+typ+"/EXPORTS")
            os.makedirs(year+"/"+quart+"/DAP_"+typ+"/PROD")
            os.makedirs(year+"/"+quart+"/MAP_"+typ+"/PROD")
            os.makedirs(year+"/"+quart+"/MGA_"+typ+"/PROD")
            os.makedirs(year+"/"+quart+"/TSP_"+typ+"/PROD")
        else:
            print year+"/"+quart+"_PROCESSED_"+typ+" exists, back up existing data, clean directory and try again"
            continue
    wb = openpyxl.load_workbook(classeur)
    sheetNames = wb.get_sheet_names()
    for sheetname in sheetNames:

        sheet = wb.get_sheet_by_name(sheetname)
        
        #GET EOR Index
        
        i=1
        while(eorPatern.match(str(sheet[get_column_letter(i)+str(1)].value)) == None ):
            i=i+1
        eor = i
        #GET ExpC Row index
        
        i=4
        cond=False
        while(expRowPatern.match(str(sheet['A'+str(i)].value)) == None):
            i=i+1
        expRow=i

        if(expPatern.match(sheet.title)== None):
            if (expPatern2.match(sheet['A1'].value) == None or typ=="ANN"):
                if typ=="ANN":
                	print "!!! : Annual production and deliveries of "+name+" neglected > "+sheet.title
                	continue
                elif (dapPtrn.match(sheet.title) != None):
                    path=year+"/"+quart+"/DAP_"+typ+"/PROD/data.csv"
                elif (mapPtrn.match(sheet.title) != None):
                    path=year+"/"+quart+"/MAP_"+typ+"/PROD/data.csv"
                elif (rckPtrn.match(sheet.title) != None):
                    path=year+"/"+quart+"/ROCK_"+typ+"/PROD/data.csv"
                elif (tspPtrn.match(sheet.title) != None):
                    path=year+"/"+quart+"/TSP_"+typ+"/PROD/data.csv"
                elif (mgaPtrn.match(sheet.title) != None):
                    path=year+"/"+quart+"/MGA_"+typ+"/PROD/data.csv"
                print (year+" "+quart+" "+"PRDCTN  > " + sheet.title)
                #Print header row
                csvrow=[]
                csvrow.append(sheet['A'+str(expRow)].value)

                for j in range(1,eor-1):
                    if j%8 == 7:
                        continue
                    if j<8:
                        strg="Production "
                    elif j<16:
                        strg="Total Deliveries "
                    elif j<24:
                        strg="Home Delieveries "
                    else:
                        strg="Exports "
                    if j%2 !=0:
                        csvrow.append(strg+str(sheet[get_column_letter(j+2)+str(expRow)].value))
                    else:
                        continue

                csvfile=open(path, 'ab')
                ifawriter = csv.writer(csvfile, quoting=csv.QUOTE_NONNUMERIC)
                ifawriter.writerow(csvrow)
                k = expRow+1
                while (eicPatern.match(str(sheet['A'+str(k)].value)) == None and eicPatern2.match(str(sheet['A'+str(k)].value)) == None):
                	if subPatern.match(str(sheet['A'+str(k)].value)) == None and str(sheet['A'+str(k)].value) != "None" and str(sheet['C'+str(k)].value) != "None":
	                    csvrow=[]
	                    for j in range(1,eor+1):
	                        if (str(sheet[get_column_letter(j)+str(k)].value)[:1] == "=" or  str(sheet['A'+str(k)].value) == "None"):
	                            continue
	                        if j%2 !=0:
	                            csvrow.append(str(sheet[get_column_letter(j)+str(k)].value))
	                        else:
	                            continue
	                    ifawriter.writerow(csvrow)
	                k=k+1

            else:
                continue

            

        else:
        	if "by Destination" in sheet['A1'].value and "BPL" not in sheet['A1'].value:
	            print (year+" "+quart+" "+"EXPORT  > " + sheet.title)
	            if (dapPtrn.match(sheet.title) != None):
	                path=year+"/"+quart+"/DAP_"+typ+"/EXPORTS/data.csv"
	            elif (mapPtrn.match(sheet.title) != None):
	                path=year+"/"+quart+"/MAP_"+typ+"/EXPORTS/data.csv"
	            elif (rckPtrn.match(sheet.title) != None):
	                path=year+"/"+quart+"/ROCK_"+typ+"/EXPORTS/data.csv"
	            elif (tspPtrn.match(sheet.title) != None):
	                path=year+"/"+quart+"/TSP_"+typ+"/EXPORTS/data.csv"
	            elif (mgaPtrn.match(sheet.title) != None):
	                path=year+"/"+quart+"/MGA_"+typ+"/EXPORTS/data.csv"
	            csvfile=open(path, 'ab')
	            ifawriter = csv.writer(csvfile, quoting=csv.QUOTE_NONNUMERIC)
	            #Get ExpC Col Index
	            
	            i=2
	            while(expColPatern.match(str(sheet[get_column_letter(i)+str(expRow)].value)) == None):
	                i=i+1
	            expCol=i+1
	            #Get exporting countries List
	            ExpList=[]
	            ExpList.append("Countries")
	            for i in range(expCol,eor-2):
	                ExpList.append(sheet[get_column_letter(i)+str(expRow)].value)
	            #Create TradeMatrix CSV
	            ifawriter.writerow(ExpList)

	            #Get ImpRow Col Index
	            
	            i=2


	            #ifawriter.writerow(['Countries']+ExpList)

	            while(impRowPatern.match(str(sheet['A'+str(i)].value)) == None):
	                i=i+1
	            impRow=i+1
	            #Get importing countries List
	            ImpList=[]

	            i=impRow
	            while(eicPatern.match(str(sheet['A'+str(i)].value)) == None):
	                csvrow=[]
	                flag=False
	                if(sheet['E'+str(i)].value == None):
	                    lastContinent = sheet['A'+str(i)].value
	                if( (sheet['E'+str(i)].value != None or sheet['A'+str(i)].value != None ) and subPatern.match(str(sheet['A'+str(i)].value)) == None ):
	                    if (sheet['A'+str(i)].value == "Various" and sheet['E'+str(i)].value != None):
	                        country= "Various " + lastContinent
	                    elif sheet['A'+str(i)].value == "Others":
	                        country= "Various Others"
	                    else:
	                        country = sheet['A'+str(i)].value
	                    csvrow.append(country)
	                    if sheet['E'+str(i)].value == None and sheet['A'+str(i)].value != None:
	                    	flag=True
	                    for j in range(2,eor-2):
	                        if(sheet[get_column_letter(j)+str(i)].value != None or flag==True):
	                        	if flag ==False:
	                        		csvrow.append(float(sheet[get_column_letter(j)+str(i)].value))
	                        	else:
	                        		csvrow.append(sheet[get_column_letter(j)+str(i)].value)

	                    ifawriter.writerow(csvrow)

	                i=i+1

